import os
import openpyxl
from openpyxl.utils import get_column_letter
import officials


def is_blank_cell(sheet, cell):
    print(sheet[cell].value)
    if sheet[cell].value != None:
        return True
    elif sheet[cell].value == 0:
        return True
    else:
        return None


def write_cell(sheet, cell, value):
    print("writing cell: %s\n" % cell)
    sheet[cell] = value


def search_name_row(sheet, person):
    name_line="B"
    row = 3
    while True:
        cell = "%s%s" % (name_line, row)
        if sheet[cell].value == person:
            return row
        elif row == 100:
            print("stop name search: %s" % person)
            break
        row += 1


def search_blank_line(sheet, row):
    line = 3
    while True:
        line_letter = get_column_letter(line)
        cell = "%s%s" % (line_letter, row)
        if is_blank_cell(sheet, cell) == None:
            return line
        elif line == 100:
            print("stop search")
            break
        line += 1


def search_person_count(person, month_result):
    for result in month_result:
        if result['name'] == person:
            print("person: %s\nsearch_person_count returned: %s" % (person, result['count']))
            return result['count']


def has_this_month_label(sheet, month):
    row = 2
    line = 2
    while True:
        checked_column = get_column_letter(line)
        cell = "%s%s" % (checked_column, row)
        result_column = "%s月" % month
        # print("checked cell: %s" % cell)
        if sheet[cell].value == result_column:
            print("value: %s" % sheet[cell].value)
            print("month hitted. skip writing")
            return False
        elif sheet[cell].value == "合計":
            print("month don't hitted, return True")
            return True
        else:
            print("value: %s" % sheet[cell].value)
            line += 1


def write_month_header(sheet, month):
    line = search_blank_line(sheet, 2)
    column = get_column_letter(line)
    cell = "%s2" % column
    sheet[cell] = "%s月" % month


def write_to_sheet(month_result, month):
    filename = 'ホメログ集計.xlsx'
    data_dir = os.path.join(os.getcwd(), 'Data', filename)
    wb = openpyxl.load_workbook(data_dir)
    sheet = wb.get_sheet_by_name('Sheet1')
    if has_this_month_label(sheet, month):
        write_month_header(sheet, month)
        for person in officials.namelist:
            row = search_name_row(sheet, person)
            print("row: %s" % row)
            line = search_blank_line(sheet, row)
            print("line: %s" % line)
            line_letter = get_column_letter(line)
            cell = "%s%s" % (line_letter, row)
            print('cell: %s' % cell)
            value = search_person_count(person, month_result)
            print("person: %s, value: %s" % (person, value))
            write_cell(sheet, cell, value)
        wb.save(data_dir)


if __name__ == '__main__':
    pass
